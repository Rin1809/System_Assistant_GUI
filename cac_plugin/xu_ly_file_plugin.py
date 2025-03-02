import os
import re
import json
import csv
import difflib
import magic
import traceback
from docx import Document
import openpyxl
from pygments import highlight
from pygments.lexers import get_lexer_by_name, guess_lexer_for_filename
from pygments.formatters import TerminalFormatter
from utils.nhat_ky import log_error, log_info
from utils import cau_hinh
from core.chat import hoi_gemini

class XuLyFilePlugin:
    def __init__(self):
        self.ten = "xử lý file nâng cao"

    def doc_file(self, duong_dan):
        """Đọc nội dung file."""
        try:
            with open(duong_dan, "r", encoding="utf-8") as f:
                noi_dung = f.read()
            return self._tao_ket_qua_xu_ly_file(duong_dan, noi_dung, "Đã đọc file")
        except Exception as e:
            log_error(f"Lỗi khi đọc file: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi đọc file: {e}")
        
    
    def xu_ly_code(self, duong_dan):
        """Đọc và định dạng code."""
        try:
            with open(duong_dan, "r", encoding="utf-8") as f:
                code = f.read()
            mime_type = magic.from_file(duong_dan, mime=True)
            try:
                lexer = get_lexer_by_name(self._chon_lexer(mime_type))
            except:
                lexer = guess_lexer_for_filename(duong_dan, code)
            code_mau = highlight(code, lexer, TerminalFormatter())
            return self._tao_ket_qua_xu_ly_file(duong_dan, code_mau, "Đã định dạng code")
        except Exception as e:
            log_error(f"Lỗi khi xử lý code: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi xử lý code: {e}")

    def xu_ly_docx(self, duong_dan):
        """Đọc nội dung file docx."""
        try:
            doc = Document(duong_dan)
            noi_dung = "\n".join([p.text for p in doc.paragraphs])
            return self._tao_ket_qua_xu_ly_file(duong_dan, noi_dung, "Đã đọc file docx")
        except Exception as e:
            log_error(f"Lỗi khi xử lý file docx: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi xử lý file docx: {e}")

    def xu_ly_xlsx(self, duong_dan):
        """Đọc nội dung file xlsx."""
        try:
            workbook = openpyxl.load_workbook(duong_dan, data_only=True)
            noi_dung = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                noi_dung += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows():
                    noi_dung += ", ".join([str(cell.value) if cell.value is not None else "" for cell in row]) + "\n"
            return self._tao_ket_qua_xu_ly_file(duong_dan, noi_dung, "Đã đọc file xlsx")
        except Exception as e:
            log_error(f"Lỗi khi xử lý file xlsx: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi xử lý file xlsx: {e}")

    def xu_ly_csv(self, duong_dan):
        """Đọc nội dung file csv."""
        try:
            with open(duong_dan, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                noi_dung = "\n".join([", ".join(row) for row in reader])
            return self._tao_ket_qua_xu_ly_file(duong_dan, noi_dung, "Đã đọc file csv")
        except Exception as e:
            log_error(f"Lỗi khi xử lý file csv: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi xử lý file csv: {e}")

    def xu_ly_json(self, duong_dan):
        """Đọc nội dung file json."""
        try:
            with open(duong_dan, "r", encoding="utf-8") as f:
                data = json.load(f)
                noi_dung = json.dumps(data, indent=4, ensure_ascii=False)
            return self._tao_ket_qua_xu_ly_file(duong_dan, noi_dung, "Đã đọc file json")
        except Exception as e:
            log_error(f"Lỗi khi xử lý file json: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi xử lý file json: {e}")

    def chinh_sua_file(self, duong_dan, lenh_chinh_sua):
        """
        Chỉnh sửa file dựa trên các lệnh:
          - thay thế "chuỗi cũ" bằng "chuỗi mới"
          - xóa "chuỗi cần xóa"
          - thêm "chuỗi cần thêm" vào cuối
        """
        try:
            with open(duong_dan, "r", encoding="utf-8") as f:
                noi_dung_cu = f.read()
            noi_dung_moi = noi_dung_cu
            
            # Thay thế
            if 'thay thế "' in lenh_chinh_sua:
                for thay_the in re.findall(r'thay thế "([^"]*)" bằng "([^"]*)"', lenh_chinh_sua):
                    noi_dung_moi = noi_dung_moi.replace(thay_the[0], thay_the[1])

            # Xóa
            if 'xóa "' in lenh_chinh_sua:
                for xoa in re.findall(r'xóa "([^"]*)"', lenh_chinh_sua):
                    noi_dung_moi = noi_dung_moi.replace(xoa, "")
            
            # Thêm
            if 'thêm "' in lenh_chinh_sua:
                for them in re.findall(r'thêm "([^"]*)" vào cuối', lenh_chinh_sua):
                    noi_dung_moi += them + "\n"

            if noi_dung_cu != noi_dung_moi:
                with open(duong_dan, "w", encoding="utf-8") as f:
                    f.write(noi_dung_moi)
                
                return self._tao_ket_qua_thay_doi(duong_dan, noi_dung_cu, noi_dung_moi, f"Đã chỉnh sửa file")
            else:
                return self._tao_ket_qua_xu_ly_file(duong_dan, None, "Không có thay đổi")

        except Exception as e:
            log_error(f"Lỗi khi chỉnh sửa file: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi chỉnh sửa file: {e}")

    def ghi_file(self, duong_dan, noi_dung):
        """Ghi nội dung vào file."""
        try:
            with open(duong_dan, "w", encoding="utf-8") as f:
                f.write(noi_dung)
            return self._tao_ket_qua_xu_ly_file(duong_dan, noi_dung, f"Đã ghi vào file")
        except Exception as e:
            log_error(f"Lỗi khi ghi vào file: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi ghi vào file: {e}")

    def xu_ly_tao_code(self, duong_dan, cau_hoi, memory):
        """Tạo code theo yêu cầu, lưu vào file và ghi vào memory"""
        prompt = f"""
        {cau_hinh.YELLOW}
        {cau_hoi}
        {cau_hinh.RESET}
        """
        try:
            phan_hoi = hoi_gemini(prompt)
            if not phan_hoi:
                return self._tao_ket_qua_loi(duong_dan, "Gemini không phản hồi.")

            code_moi = re.sub(r"```(python)?", "", phan_hoi).strip() # bo language tag neu co
            if code_moi:
                self.ghi_file(duong_dan, code_moi) # luu vao file
                log_info(f"Đã tạo code và ghi vào file {duong_dan}")
                self._luu_memory_tao_code(duong_dan, code_moi, cau_hoi, memory)
                return self._tao_ket_qua_xu_ly_file(duong_dan, code_moi, f"Đã tạo code và ghi vào file {duong_dan}.")
            else:
                log_info(f"Không tạo được code cho file '{duong_dan}'")
                return self._tao_ket_qua_xu_ly_file(duong_dan, None, f"Không tạo được code như yêu cầu.")
        except Exception as e:
            log_error(f"Lỗi khi tạo code: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi tạo code: {e}")
            
    def fix_code(self, duong_dan, cau_hoi, memory):
        """Sửa lỗi code trong file."""
        try:
            with open(duong_dan, "r", encoding="utf-8") as f:
                code_cu = f.read()

            prompt = f"""
                {cau_hinh.YELLOW}
                Đây là code:
                ```
                {code_cu}
                ```
                {cau_hoi} Sửa lỗi trong code trên, không cần giải thích gì cả, chỉ cần đưa ra code đã fix
                {cau_hinh.RESET}
            """

            phan_hoi = hoi_gemini(prompt)
            if phan_hoi:
                code_moi = re.sub(r"```(python)?", "", phan_hoi).strip()
                if code_moi and code_moi != code_cu:
                    with open(duong_dan, "w", encoding="utf-8") as f:
                        f.write(code_moi)
                    
                    self._luu_memory_fix_code(duong_dan, code_cu, code_moi, cau_hoi, memory) # Thêm hàm này để lưu memory
                    return self._tao_ket_qua_thay_doi(duong_dan, code_cu, code_moi, "Đã sửa lỗi và cập nhật code")
                else:
                    return self._tao_ket_qua_loi(duong_dan, "Không thể sửa lỗi code")
            else:
                return self._tao_ket_qua_loi(duong_dan, "Gemini không phản hồi")
        except Exception as e:
            log_error(f"Lỗi khi sửa lỗi code: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi sửa lỗi code: {e}")

    def nang_cap_code(self, duong_dan, cau_hoi, memory):
        """Nâng cấp code trong file."""
        try:
            with open(duong_dan, "r", encoding="utf-8") as f:
                code_cu = f.read()

            prompt = f"""
                {cau_hinh.YELLOW}
                Đây là code:
                ```
                {code_cu}
                ```
                {cau_hoi} Hãy nâng cấp code này, không cần giải thích, chỉ đưa ra code sau khi đã nâng cấp
                {cau_hinh.RESET}
            """

            phan_hoi = hoi_gemini(prompt)
            if phan_hoi:
                code_moi = re.sub(r"```(python)?", "", phan_hoi).strip()
                if code_moi and code_moi != code_cu:
                    with open(duong_dan, "w", encoding="utf-8") as f:
                        f.write(code_moi)
                    
                    self._luu_memory_nang_cap_code(duong_dan, code_cu, code_moi, cau_hoi, memory)
                    return self._tao_ket_qua_thay_doi(duong_dan, code_cu, code_moi, "Đã nâng cấp code")
                else:
                    return self._tao_ket_qua_loi(duong_dan, "Không có thay đổi trên code")
            else:
                return self._tao_ket_qua_loi(duong_dan, "Gemini không phản hồi")
        except Exception as e:
            log_error(f"Lỗi khi nâng cấp code: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi nâng cấp code: {e}")

    def _luu_memory_fix_code(self, duong_dan, code_cu, code_moi, cau_hoi, memory):
        """Lưu thông tin sửa code vào memory"""
        memory.append({
            "type": "xu_ly_file",
            "filepath": duong_dan,
            "nội dung trước": code_cu,
            "nội dung sau": code_moi,
            "câu hỏi": cau_hoi,
            "message": f"Đã sửa code và ghi vào file {duong_dan}.",
        })

    def _luu_memory_nang_cap_code(self, duong_dan, code_cu, code_moi, cau_hoi, memory):
        """Lưu thông tin nâng cấp code vào memory"""
        memory.append({
            "type": "xu_ly_file",
            "filepath": duong_dan,
            "nội dung trước": code_cu,
            "nội dung sau": code_moi,
            "câu hỏi": cau_hoi,
            "message": f"Đã nâng cấp code và ghi vào file {duong_dan}.",
        })

    
    def _luu_memory_tao_code(self, duong_dan, code_moi, cau_hoi, memory):
        """Lưu thông tin tạo code vào memory"""
        memory.append({
            "type": "xu_ly_file",
            "filepath": duong_dan,
            "nội dung": code_moi,
            "câu hỏi": cau_hoi,
            "message": f"Đã tạo code và ghi vào file {duong_dan}.",
        })

    def xu_ly_file(self, duong_dan, cau_hoi, memory):
        """Hàm xử lý chính cho các file."""
        if not os.path.exists(duong_dan):
            return self._tao_ket_qua_loi(duong_dan, "File không tồn tại")

        mime_type = magic.from_file(duong_dan, mime=True)

        try:
            if "text" in mime_type or "application/json" in mime_type:
                if "chỉnh sửa" in cau_hoi or "sửa" in cau_hoi:
                    return self.chinh_sua_file(duong_dan, cau_hoi)
                elif re.search(r'viết nội dung\s*"(.*?)"', cau_hoi, re.IGNORECASE):
                    noi_dung = re.search(r'viết nội dung\s*"(.*?)"', cau_hoi, re.IGNORECASE).group(1)
                    return self.ghi_file(duong_dan, noi_dung)
                elif re.search(r"(viết cho tôi một mã python|tạo một mã python)", cau_hoi, re.IGNORECASE):
                    return self.xu_ly_tao_code(duong_dan, cau_hoi, memory)
                elif "fix" in cau_hoi.lower() and 'python' in mime_type:
                    return self.fix_code(
                        duong_dan, cau_hoi, memory
                    )
                elif "nâng cấp" in cau_hoi.lower() and 'python' in mime_type:
                    return self.nang_cap_code(
                        duong_dan, cau_hoi, memory
                    )
                else:
                    return self.doc_file(duong_dan)
            else:
                return self._tao_ket_qua_loi(duong_dan, f"Loại file không được hỗ trợ: {mime_type}")
        except Exception as e:
            log_error(f"Lỗi khi xử lý file: {e}", detail=traceback.format_exc())
            return self._tao_ket_qua_loi(duong_dan, f"Lỗi khi xử lý file: {e}")
    
    def _chon_lexer(self, mime_type):
        """Chọn lexer phù hợp cho việc highlight code."""
        if 'python' in mime_type:
            return "python"
        elif 'javascript' in mime_type:
            return "javascript"
        elif 'java' in mime_type:
            return "java"
        elif 'cpp' in mime_type or 'x-c++src' in mime_type:
            return "cpp"
        elif 'html' in mime_type:
            return "html"
        elif 'css' in mime_type:
            return "css"
        else:
            return "text"

    def _tao_ket_qua_xu_ly_file(self, filepath, content, message):
        """Tạo kết quả trả về cho các tác vụ xử lý file."""
        return {
            "type": "xu_ly_file",
            "filepath": filepath,
            "content": content,
            "message": message,
        }
    
    def _tao_ket_qua_thay_doi(self, filepath, code_cu, code_moi, message):
        """Tạo kết quả trả về cho các tác vụ xử lý file có thay đổi."""
        return {
            "type": "thay_doi",
            "filepath": filepath,
            "code_cu": code_cu,
            "code_moi": code_moi,
            "message": message,
        }

    def _tao_ket_qua_loi(self, filepath, message):
        """Tạo kết quả trả về khi có lỗi xảy ra."""
        return {
            "type": "error",
            "filepath": filepath,
            "message": message
        }